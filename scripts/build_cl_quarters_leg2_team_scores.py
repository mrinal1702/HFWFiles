from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path
import sys
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\trive\HFWFiles")
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from procedures.best_xi import compute_best_xi, load_master_player_list

TABLES_DIR = ROOT / "Matches_Raw" / "CL Quarters Leg2 Tables"
OWNERSHIP_TSV = TABLES_DIR / "master_player_list - master_player_list.tsv"
MATCHES_DIR = ROOT / "Matches_Raw" / "CL_Quarters_Leg2"
MASTER_PLAYER_CSV = ROOT / "Player_List" / "master_player_list.csv"

MERGED_SCORES_OUT = TABLES_DIR / "CL_Quarters_Leg2_AllPlayerScores.csv"
TEAM_TOTALS_OUT = TABLES_DIR / "CL_Quarters_Leg2_TeamTotals.csv"
TEAM_BREAKDOWN_OUT = TABLES_DIR / "CL_Quarters_Leg2_TeamBreakdown_BestXI.csv"
WORKBOOK_OUT = TABLES_DIR / "CL_Quarters_Leg2_Scoring.xlsx"

FINAL_POINTS_GLOB = "*FinalPoints.csv"


def read_delimited_rows(path: Path, delimiter: str) -> list[dict[str, str]]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                return list(csv.DictReader(f, delimiter=delimiter))
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("unknown", b"", 0, 1, f"Unable to decode {path}")


def normalize_owner(owner: str | None) -> str:
    raw = (owner or "").strip()
    if not raw or raw.upper() == "N/A":
        return ""
    if raw in {"Shah Brothers", "Shah Bros"}:
        return "Shah Bros"
    return raw


def load_ownership_rows(path: Path) -> list[dict[str, Any]]:
    rows = read_delimited_rows(path, "\t")

    # User override: Lenglet belongs to Shah Bros.
    # User override: Johnny Cardoso remains owned for overall score, but is excluded from best XI only.
    for row in rows:
        try:
            pid = int(row.get("player_id") or -1)
        except ValueError:
            continue
        if pid == 488412:
            row["Team"] = "Shah Bros"
        elif row.get("Team") in {"Shah Brothers", "Shah Bros"}:
            row["Team"] = "Shah Bros"
    return rows


def load_merged_scores(matches_dir: Path) -> tuple[list[dict[str, Any]], dict[int, int]]:
    merged_rows: list[dict[str, Any]] = []
    score_by_id: dict[int, int] = {}
    for csv_path in sorted(matches_dir.glob(FINAL_POINTS_GLOB)):
        rows = read_delimited_rows(csv_path, ",")
        for row in rows:
            raw_id = (row.get("player_id") or "").strip()
            if not raw_id:
                continue
            try:
                pid = int(raw_id)
            except ValueError:
                continue
            try:
                final_score = int(float(row.get("final_score") or 0))
            except ValueError:
                final_score = 0
            merged_rows.append(
                {
                    "source_file": csv_path.name,
                    "player_name": (row.get("player_name") or "").strip(),
                    "player_id": pid,
                    "team_name": (row.get("team_name") or "").strip(),
                    "position": (row.get("position") or "").strip(),
                    "stats_score": row.get("stats_score") or "",
                    "endowment_score": row.get("endowment_score") or "",
                    "final_score": final_score,
                }
            )
            score_by_id[pid] = final_score
    merged_rows.sort(key=lambda r: (r["team_name"], r["position"], -int(r["final_score"]), int(r["player_id"])))
    return merged_rows, score_by_id


def build_owner_squads(
    ownership_rows: list[dict[str, Any]],
    master_players: dict[int, dict[str, Any]],
) -> tuple[dict[str, list[int]], dict[str, set[int]]]:
    owner_outfield_ids: dict[str, list[int]] = defaultdict(list)
    owner_keeper_team_ids: dict[str, set[int]] = defaultdict(set)

    for row in ownership_rows:
        owner = normalize_owner(row.get("Team"))
        if not owner:
            continue
        try:
            player_id = int(row.get("player_id") or -1)
            team_id = int(row.get("team_id") or -1)
        except ValueError:
            continue
        position = (row.get("position") or "").strip().lower()
        if position == "goalkeeper":
            if team_id >= 0:
                owner_keeper_team_ids[owner].add(team_id)
            continue
        owner_outfield_ids[owner].append(player_id)

    owner_squad_ids: dict[str, list[int]] = {}
    for owner in sorted(set(owner_outfield_ids) | set(owner_keeper_team_ids)):
        squad_ids = list(dict.fromkeys(owner_outfield_ids.get(owner, [])))
        keeper_team_ids = owner_keeper_team_ids.get(owner, set())
        if keeper_team_ids:
            for pid, meta in master_players.items():
                if str(meta.get("position") or "").strip().lower() != "goalkeeper":
                    continue
                team_id = meta.get("team_id")
                if team_id in keeper_team_ids:
                    squad_ids.append(pid)
        owner_squad_ids[owner] = list(dict.fromkeys(squad_ids))

    return owner_squad_ids, owner_keeper_team_ids


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def autosize_columns(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 40)


def safe_sheet_name(name: str) -> str:
    bad = set(r'[]:*?/\\')
    cleaned = "".join("_" if c in bad else c for c in name).strip()
    return cleaned[:31] or "Sheet"


def write_owner_sheet(ws, owner: str, overall_total: int, best_result: Any, selected_rows: list[dict[str, Any]], left_out_rows: list[dict[str, Any]]) -> None:
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    ws["A1"] = owner
    ws["A1"].font = title_font
    ws["A3"] = "Formation"
    ws["B3"] = str(best_result.formation_label)
    ws["A4"] = "Best XI Total"
    ws["B4"] = int(best_result.total_points)
    ws["A5"] = "Overall Total"
    ws["B5"] = int(overall_total)
    for cell in ("A3", "A4", "A5"):
        ws[cell].font = bold

    current_row = 7
    ws.cell(current_row, 1, "Best XI")
    ws.cell(current_row, 1).font = bold
    current_row += 1
    ws.cell(current_row, 1, "Line")
    ws.cell(current_row, 2, "Player")
    ws.cell(current_row, 3, "Club")
    ws.cell(current_row, 4, "Score")
    for col in range(1, 5):
        ws.cell(current_row, col).font = bold
    current_row += 1

    line_order = {"GK": 0, "D": 1, "M": 2, "F": 3}
    selected_rows_sorted = sorted(selected_rows, key=lambda r: (line_order.get(r["line_role"], 9), -int(r["score"]), str(r["player_name"])))
    for row in selected_rows_sorted:
        ws.cell(current_row, 1, row["line_role"])
        ws.cell(current_row, 2, row["player_name"])
        ws.cell(current_row, 3, row["club_team_name"])
        ws.cell(current_row, 4, int(row["score"]))
        current_row += 1

    ws.cell(current_row, 1, "Total")
    ws.cell(current_row, 4, int(best_result.total_points))
    ws.cell(current_row, 1).font = bold
    ws.cell(current_row, 4).font = bold
    current_row += 2

    ws.cell(current_row, 1, "Not Selected")
    ws.cell(current_row, 1).font = bold
    current_row += 1
    ws.cell(current_row, 1, "Player")
    ws.cell(current_row, 2, "Club")
    ws.cell(current_row, 3, "Listed Position")
    ws.cell(current_row, 4, "Score")
    ws.cell(current_row, 5, "Note")
    for col in range(1, 6):
        ws.cell(current_row, col).font = bold
    current_row += 1

    left_out_sorted = sorted(left_out_rows, key=lambda r: (-int(r["score"]), str(r["player_name"])))
    for row in left_out_sorted:
        ws.cell(current_row, 1, row["player_name"])
        ws.cell(current_row, 2, row["club_team_name"])
        ws.cell(current_row, 3, row["listed_position"])
        ws.cell(current_row, 4, int(row["score"]))
        ws.cell(current_row, 5, row["note"])
        current_row += 1

    autosize_columns(ws)
    ws.freeze_panes = "A8"


def write_workbook(
    team_total_rows: list[dict[str, Any]],
    merged_rows: list[dict[str, Any]],
    owner_presentations: dict[str, dict[str, Any]],
) -> None:
    wb = Workbook()
    ws_totals = wb.active
    ws_totals.title = "TeamTotals"

    totals_headers = ["Owner", "Best XI Total", "Formation", "Empty Outfield Slots", "Best XI Players Selected"]
    ws_totals.append(totals_headers)
    for cell in ws_totals[1]:
        cell.font = Font(bold=True)
    for row in team_total_rows:
        ws_totals.append(
            [
                row["owner"],
                int(row["best_xi_total"]),
                str(row["formation"]),
                int(row["empty_outfield_slots"]),
                int(row["best_xi_players_selected"]),
            ]
        )
    autosize_columns(ws_totals)
    ws_totals.freeze_panes = "A2"

    ws_scores = wb.create_sheet("AllPlayerScores")
    score_headers = ["Source File", "Player Name", "Team Name", "Position", "Stats Score", "Endowment Score", "Final Score"]
    ws_scores.append(score_headers)
    for cell in ws_scores[1]:
        cell.font = Font(bold=True)
    for row in merged_rows:
        ws_scores.append(
            [
                row["source_file"],
                row["player_name"],
                row["team_name"],
                row["position"],
                row["stats_score"],
                row["endowment_score"],
                int(row["final_score"]),
            ]
        )
    autosize_columns(ws_scores)
    ws_scores.freeze_panes = "A2"

    for owner, payload in owner_presentations.items():
        ws_owner = wb.create_sheet(safe_sheet_name(owner))
        write_owner_sheet(
            ws_owner,
            owner=owner,
            overall_total=int(payload["overall_total"]),
            best_result=payload["best_result"],
            selected_rows=payload["selected_rows"],
            left_out_rows=payload["left_out_rows"],
        )

    wb.save(WORKBOOK_OUT)


def main() -> None:
    master_players = load_master_player_list(MASTER_PLAYER_CSV)
    ownership_rows = load_ownership_rows(OWNERSHIP_TSV)
    merged_rows, score_by_id = load_merged_scores(MATCHES_DIR)
    match_json_paths = sorted(MATCHES_DIR.glob("*.json"))

    write_csv(
        MERGED_SCORES_OUT,
        [
            "source_file",
            "player_name",
            "player_id",
            "team_name",
            "position",
            "stats_score",
            "endowment_score",
            "final_score",
        ],
        merged_rows,
    )

    owner_squad_ids, owner_keeper_team_ids = build_owner_squads(ownership_rows, master_players)

    breakdown_rows: list[dict[str, Any]] = []
    team_total_rows: list[dict[str, Any]] = []
    owner_presentations: dict[str, dict[str, Any]] = {}

    grouped_by_owner: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in ownership_rows:
        owner = normalize_owner(row.get("Team"))
        if not owner:
            continue
        grouped_by_owner[owner].append(row)

    for owner in sorted(grouped_by_owner):
        rows = grouped_by_owner[owner]
        overall_total = 0
        picked_ids_for_overall: set[int] = set()
        owned_detail_rows: list[dict[str, Any]] = []

        # Count keeper-unit score once per owned goalkeeper team.
        for keeper_team_id in sorted(owner_keeper_team_ids.get(owner, set())):
            score = int(score_by_id.get(keeper_team_id, 0))
            overall_total += score
            owned_detail_rows.append(
                {
                    "owner": owner,
                    "section": "owned_players",
                    "formation": "",
                    "best_xi_total": "",
                    "overall_total": "",
                    "line_role": "",
                    "included_in_best_xi": "",
                    "player_id": keeper_team_id,
                    "player_name": f"{next((r.get('team_name') for r in rows if r.get('position') == 'Goalkeeper' and str(r.get('team_id')) == str(keeper_team_id)), 'Team')} Keepers",
                    "club_team_name": next((r.get("team_name") for r in rows if str(r.get("team_id")) == str(keeper_team_id)), ""),
                    "listed_position": "Goalkeeper",
                    "score": score,
                    "note": "Keeper unit counted once for owned goalkeeper team.",
                }
            )

        for row in rows:
            position = (row.get("position") or "").strip()
            if position.lower() == "goalkeeper":
                continue
            try:
                player_id = int(row.get("player_id") or -1)
            except ValueError:
                continue
            if player_id in picked_ids_for_overall:
                continue
            picked_ids_for_overall.add(player_id)
            score = int(score_by_id.get(player_id, 0))
            overall_total += score
            owned_detail_rows.append(
                {
                    "owner": owner,
                    "section": "owned_players",
                    "formation": "",
                    "best_xi_total": "",
                    "overall_total": "",
                    "line_role": "",
                    "included_in_best_xi": "",
                    "player_id": player_id,
                    "player_name": (row.get("player_name") or "").strip(),
                    "club_team_name": (row.get("team_name") or "").strip(),
                    "listed_position": position,
                    "score": score,
                    "note": "",
                }
            )

        best_xi_squad_ids = list(owner_squad_ids.get(owner, []))
        if owner == "Dosa XI":
            best_xi_squad_ids = [pid for pid in best_xi_squad_ids if pid != 1173678]

        best_result = compute_best_xi(
            auction_user_id=owner,
            gw_id=2,
            squad_player_ids=best_xi_squad_ids,
            gw_scores_by_player=score_by_id,
            master_players=master_players,
            match_json_paths=match_json_paths,
        )

        team_total_rows.append(
            {
                "owner": owner,
                "best_xi_total": best_result.total_points,
                "formation": best_result.formation_label,
                "empty_outfield_slots": best_result.empty_outfield_slots,
                "best_xi_players_selected": len(best_result.outfield) + (1 if best_result.goalkeeper_id is not None else 0),
            }
        )

        pick_by_id = {pick.player_id: pick for pick in best_result.outfield}
        selected_rows_for_owner: list[dict[str, Any]] = []
        left_out_rows_for_owner: list[dict[str, Any]] = []

        if best_result.goalkeeper_id is not None:
            keeper_team_name = master_players.get(best_result.goalkeeper_id, {}).get("team_name", "")
            selected_rows_for_owner.append(
                {
                    "line_role": "GK",
                    "player_name": best_result.goalkeeper_name,
                    "club_team_name": keeper_team_name,
                    "listed_position": "Goalkeeper",
                    "score": best_result.goalkeeper_score,
                    "note": "Selected goalkeeper for best XI.",
                }
            )

        for detail in owned_detail_rows:
            if detail["listed_position"] == "Goalkeeper":
                detail["formation"] = best_result.formation_label
                detail["best_xi_total"] = best_result.total_points
                detail["overall_total"] = overall_total
                detail["line_role"] = "GK" if best_result.goalkeeper_id else ""
                detail["included_in_best_xi"] = "yes" if best_result.goalkeeper_id else "no"
                breakdown_rows.append(detail)
                continue

            player_id = int(detail["player_id"])
            pick = pick_by_id.get(player_id)
            detail["formation"] = best_result.formation_label
            detail["best_xi_total"] = best_result.total_points
            detail["overall_total"] = overall_total
            detail["line_role"] = pick.role if pick else ""
            detail["included_in_best_xi"] = "yes" if pick else "no"
            if owner == "Dosa XI" and player_id == 1173678:
                detail["note"] = "Excluded from best XI per user instruction."
            elif pick:
                detail["note"] = "Selected in best XI."
            else:
                detail["note"] = "Left out of best XI."
            breakdown_rows.append(detail)
            presentation_row = {
                "line_role": detail["line_role"],
                "player_name": detail["player_name"],
                "club_team_name": detail["club_team_name"],
                "listed_position": detail["listed_position"],
                "score": detail["score"],
                "note": detail["note"],
            }
            if pick:
                selected_rows_for_owner.append(presentation_row)
            else:
                left_out_rows_for_owner.append(presentation_row)

        owner_presentations[owner] = {
            "overall_total": overall_total,
            "best_result": best_result,
            "selected_rows": selected_rows_for_owner,
            "left_out_rows": left_out_rows_for_owner,
        }

    team_total_rows.sort(key=lambda r: (-int(r["best_xi_total"]), r["owner"]))
    breakdown_rows.sort(key=lambda r: (r["owner"], 0 if r["section"] == "best_xi_summary" else 1, r["included_in_best_xi"] != "yes", -int(r["score"]) if str(r["score"]).strip() else 0, str(r["player_name"])))

    write_csv(
        TEAM_TOTALS_OUT,
        [
            "owner",
            "best_xi_total",
            "formation",
            "empty_outfield_slots",
            "best_xi_players_selected",
        ],
        team_total_rows,
    )
    write_csv(
        TEAM_BREAKDOWN_OUT,
        [
            "owner",
            "section",
            "formation",
            "best_xi_total",
            "overall_total",
            "line_role",
            "included_in_best_xi",
            "player_id",
            "player_name",
            "club_team_name",
            "listed_position",
            "score",
            "note",
        ],
        breakdown_rows,
    )
    write_workbook(team_total_rows, merged_rows, owner_presentations)

    print(f"Wrote merged player scores: {MERGED_SCORES_OUT}")
    print(f"Wrote team totals: {TEAM_TOTALS_OUT}")
    print(f"Wrote team breakdown: {TEAM_BREAKDOWN_OUT}")
    print(f"Wrote workbook: {WORKBOOK_OUT}")


if __name__ == "__main__":
    main()
